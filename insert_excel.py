import openpyxl

my_wb = openpyxl.Workbook()
my_sheet = my_wb.active

c1 = my_sheet.cell(row=1, column=1)
c1.value = "Aadrika"

c2 = my_sheet.cell(row=1, column=2)
c2.value = "Adwaita"

c3 = my_sheet['A2']
c3.value = "Satyajit"

c4 = my_sheet['B2']
c4.value = "Bivas"

my_wb.save(r"C:\Users\Pragyandeep\Desktop\Book1.xlsx")

my_wb.create_sheet(index=1, title="new sheet")
my_wb.save(r"C:\Users\Pragyandeep\Desktop\Book1.xlsx")

my_path = r"C:\Users\Pragyandeep\Desktop\Book1.xlsx"
my_wb_obj = openpyxl.load_workbook(my_path)
my_sheet_obj = my_wb_obj.active
print(my_sheet_obj.max_row)

wb_obj = openpyxl.load_workbook(my_path)
my_sheet_obj = wb_obj.active
my_cell_obj = my_sheet_obj.cell(row=1, column=1)
print(my_cell_obj.value)

print(my_sheet_obj.max_column)

my_wb_obj = openpyxl.load_workbook(my_path)
my_sheet_obj = my_wb_obj.active
my_max_col = my_sheet_obj.max_column
for i in range(1, my_max_col + 1):
   my_cell_obj = my_sheet_obj.cell(row = 1, column = i)
   print(my_cell_obj.value)

my_row = my_sheet_obj.max_row
for i in range(1, my_row + 1):
   cell_obj = my_sheet_obj.cell(row = i, column = 1)
   print(cell_obj.value)

for i in range(1, my_max_col + 1):
   cell_obj = my_sheet_obj.cell(row = 2, column = i)
   print(cell_obj.value, end = " ")
